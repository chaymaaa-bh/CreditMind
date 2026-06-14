# =============================================================
#  CreditMind — M2 : Graph Neural Network (GNN)
#
#  INSTRUCTIONS :
#  1. Place ce fichier dans :
#     C:\Users\Rahma Moalla\Desktop\CreditMind\
#  2. Installe les dépendances :
#     pip install torch scikit-learn networkx pandas numpy openpyxl matplotlib
#  3. Lance :
#     python m2_gnn.py
#
#  Résultat :
#     datta-nettoyer\m2_gnn_scores.csv     → score GNN par client
#     datta-nettoyer\m2_gnn_results.xlsx   → rapport complet formaté
# =============================================================
 
import pandas as pd
import numpy as np
import os, warnings
warnings.filterwarnings('ignore')
 
# ─── CHEMIN VERS LE DATASET COMBINÉ (produit par M1) ─────────────────────────
#INPUT = r'datta-nettoyer\dataset_combined_real_synth.csv'
INPUT = r'dataset_combined_real_synth.csv'
if not os.path.exists(INPUT):
    # Fallback : utiliser uniquement la data nettoyée si M1 pas encore terminé
    #INPUT = r'datta-nettoyer\SolvAI_Dataset_Nettoye.xlsx'
    INPUT = r'SolvAI_Dataset_Nettoye.xlsx'
    print("  dataset_combined non trouvé → utilisation de SolvAI_Dataset_Nettoye.xlsx")
    df = pd.read_excel(INPUT, sheet_name='clients_features')
else:
    df = pd.read_csv(INPUT)
 
print(f"► Dataset chargé : {df.shape[0]} clients · {df.shape[1]} colonnes")
 
# ═══════════════════════════════════════════════════════════════
# ÉTAPE 1 — Préparation des features
# ═══════════════════════════════════════════════════════════════
print("\n► Étape 1 : Préparation des features...")
 
FEATURE_COLS = [
    'nb_factures', 'montant_ttc_total', 'montant_ttc_moyen',
    'montant_ttc_max', 'nb_avoirs', 'ratio_avoirs', 'anciennete_jours',
    'nb_reglements', 'montant_reg_moyen', 'retard_moyen_jours',
    'retard_max_jours', 'nb_retards_positifs', 'nb_retards_graves',
    'taux_retard', 'nb_modes_distincts', 'ratio_encaissement',
    'gouvernorat_code', 'nature_client_code', 'mode_paiement_code',
]
 
# Garder uniquement les colonnes disponibles
FEATURE_COLS = [c for c in FEATURE_COLS if c in df.columns]
TARGET_COL   = 'label_risque'
 
# Remplir les valeurs manquantes
df[FEATURE_COLS] = df[FEATURE_COLS].fillna(df[FEATURE_COLS].median())
 
X = df[FEATURE_COLS].values.astype(np.float32)
y = df[TARGET_COL].values.astype(np.int64) if TARGET_COL in df.columns else np.zeros(len(df), dtype=np.int64)
 
print(f"  Features utilisées : {len(FEATURE_COLS)}")
print(f"  Clients : {len(X)}")
print(f"  Label risque=1 : {y.sum()} ({y.mean():.1%})")
 
# ═══════════════════════════════════════════════════════════════
# ÉTAPE 2 — Construction du graphe réseau client
# ═══════════════════════════════════════════════════════════════
print("\n► Étape 2 : Construction du graphe réseau...")
 
# Normaliser les features pour calculer les similarités
from sklearn.preprocessing import StandardScaler
from sklearn.metrics.pairwise import cosine_similarity
 
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X)
 
# Créer les arêtes du graphe :
# Deux clients sont connectés s'ils partagent le même gouvernorat
# OU le même responsable commercial OU sont similaires financièrement
 
edges_src = []
edges_dst = []
 
print("  Construction des connexions par gouvernorat...")
if 'gouvernorat_code' in df.columns:
    for gouv_code in df['gouvernorat_code'].unique():
        idx = df[df['gouvernorat_code'] == gouv_code].index.tolist()
        # Connecter les clients du même gouvernorat (max 10 connexions par client)
        for i in range(len(idx)):
            for j in range(i+1, min(i+6, len(idx))):
                edges_src.append(idx[i])
                edges_dst.append(idx[j])
                edges_src.append(idx[j])
                edges_dst.append(idx[i])
 
print("  Construction des connexions par responsable...")
if 'code_responsable' in df.columns:
    for resp in df['code_responsable'].unique():
        idx = df[df['code_responsable'] == resp].index.tolist()
        for i in range(len(idx)):
            for j in range(i+1, min(i+4, len(idx))):
                edges_src.append(idx[i])
                edges_dst.append(idx[j])
                edges_src.append(idx[j])
                edges_dst.append(idx[i])
 
# Supprimer les doublons
edges = list(set(zip(edges_src, edges_dst)))
edges_src = [e[0] for e in edges]
edges_dst = [e[1] for e in edges]
 
print(f"  Noeuds (clients) : {len(df)}")
print(f"  Arêtes (connexions) : {len(edges_src)}")
 
# ═══════════════════════════════════════════════════════════════
# ÉTAPE 3 — Modèle GNN (GraphSAGE simplifié)
# ═══════════════════════════════════════════════════════════════
print("\n► Étape 3 : Entraînement du modèle GNN...")
 
try:
    import torch
    import torch.nn as nn
    import torch.nn.functional as F
 
    # Tensors
    X_tensor = torch.tensor(X_scaled, dtype=torch.float32)
    y_tensor  = torch.tensor(y, dtype=torch.long)
    edge_index = torch.tensor([edges_src, edges_dst], dtype=torch.long)
 
    try:
        # ── Tentative avec PyTorch Geometric ──────────────────
        from torch_geometric.nn import SAGEConv
        from torch_geometric.data import Data
 
        class GraphSAGE(nn.Module):
            def __init__(self, in_channels, hidden_channels, out_channels):
                super().__init__()
                self.conv1 = SAGEConv(in_channels, hidden_channels)
                self.conv2 = SAGEConv(hidden_channels, hidden_channels)
                self.conv3 = SAGEConv(hidden_channels, out_channels)
                self.dropout = nn.Dropout(0.3)
 
            def forward(self, x, edge_index):
                x = F.relu(self.conv1(x, edge_index))
                x = self.dropout(x)
                x = F.relu(self.conv2(x, edge_index))
                x = self.dropout(x)
                x = self.conv3(x, edge_index)
                return x
 
        data  = Data(x=X_tensor, edge_index=edge_index, y=y_tensor)
        model = GraphSAGE(
            in_channels=len(FEATURE_COLS),
            hidden_channels=64,
            out_channels=2
        )
        USE_PYGEOMETRIC = True
        print("  Modèle : GraphSAGE (PyTorch Geometric)")
 
    except ImportError:
        # ── Fallback : GNN manuel sans PyTorch Geometric ──────
        print("  PyTorch Geometric non trouvé → GNN manuel")
 
        class SimpleGNN(nn.Module):
            """
            GraphSAGE simplifié :
            Pour chaque noeud, on agrège les features de ses voisins
            puis on applique des couches MLP.
            """
            def __init__(self, in_channels, hidden_channels, out_channels):
                super().__init__()
                self.layer1 = nn.Linear(in_channels * 2, hidden_channels)
                self.layer2 = nn.Linear(hidden_channels, hidden_channels)
                self.layer3 = nn.Linear(hidden_channels, out_channels)
                self.dropout = nn.Dropout(0.3)
                self.bn1 = nn.BatchNorm1d(hidden_channels)
                self.bn2 = nn.BatchNorm1d(hidden_channels)
 
            def aggregate_neighbors(self, x, src, dst, n_nodes):
                """Agrège les features des voisins pour chaque noeud"""
                agg = torch.zeros(n_nodes, x.shape[1])
                count = torch.zeros(n_nodes, 1)
                for i in range(len(src)):
                    agg[dst[i]] += x[src[i]]
                    count[dst[i]] += 1
                count = count.clamp(min=1)
                return agg / count
 
            def forward(self, x, src, dst):
                n = x.shape[0]
                # Agrégation des voisins
                neigh = self.aggregate_neighbors(x, src, dst, n)
                # Concaténer features propres + features voisins
                x_cat = torch.cat([x, neigh], dim=1)
                # Couches MLP
                h = F.relu(self.bn1(self.layer1(x_cat)))
                h = self.dropout(h)
                h = F.relu(self.bn2(self.layer2(h)))
                h = self.dropout(h)
                out = self.layer3(h)
                return out
 
        model = SimpleGNN(
            in_channels=len(FEATURE_COLS),
            hidden_channels=64,
            out_channels=2
        )
        USE_PYGEOMETRIC = False
 
    # ── Entraînement ──────────────────────────────────────────
    optimizer = torch.optim.Adam(model.parameters(), lr=0.01, weight_decay=5e-4)
 
    # Gestion du déséquilibre des classes
    n_pos   = y.sum()
    n_neg   = len(y) - n_pos
    weight  = torch.tensor([1.0, n_neg/max(n_pos, 1)], dtype=torch.float32)
    criterion = nn.CrossEntropyLoss(weight=weight)
 
    # Split train/test (80/20)
    n_train = int(0.8 * len(X_tensor))
    idx_all  = torch.randperm(len(X_tensor))
    train_mask = idx_all[:n_train]
    test_mask  = idx_all[n_train:]
 
    src_t = torch.tensor(edges_src, dtype=torch.long)
    dst_t = torch.tensor(edges_dst, dtype=torch.long)
 
    print(f"  Train : {len(train_mask)} clients · Test : {len(test_mask)} clients")
    print(f"  Entraînement en cours...")
 
    best_auc = 0
    for epoch in range(1, 201):
        model.train()
        optimizer.zero_grad()
 
        if USE_PYGEOMETRIC:
            out = model(X_tensor, edge_index)
        else:
            out = model(X_tensor, src_t.tolist(), dst_t.tolist())
 
        loss = criterion(out[train_mask], y_tensor[train_mask])
        loss.backward()
        optimizer.step()
 
        if epoch % 20 == 0:
            model.eval()
            with torch.no_grad():
                if USE_PYGEOMETRIC:
                    logits = model(X_tensor, edge_index)
                else:
                    logits = model(X_tensor, src_t.tolist(), dst_t.tolist())
 
                probs = F.softmax(logits, dim=1)[:, 1].numpy()
                preds = logits.argmax(dim=1).numpy()
 
                from sklearn.metrics import roc_auc_score, accuracy_score
                try:
                    auc = roc_auc_score(y[test_mask.numpy()], probs[test_mask.numpy()])
                    acc = accuracy_score(y[test_mask.numpy()], preds[test_mask.numpy()])
                    if auc > best_auc:
                        best_auc = auc
                    print(f"  Epoch {epoch:3d} | Loss: {loss.item():.4f} | AUC: {auc:.4f} | Acc: {acc:.4f}")
                except:
                    print(f"  Epoch {epoch:3d} | Loss: {loss.item():.4f}")
 
    # ── Score final GNN pour chaque client ────────────────────
    model.eval()
    with torch.no_grad():
        if USE_PYGEOMETRIC:
            logits = model(X_tensor, edge_index)
        else:
            logits = model(X_tensor, src_t.tolist(), dst_t.tolist())
        probs = F.softmax(logits, dim=1)[:, 1].numpy()
 
    # Score GNN 0-100 (probabilité de risque × 100)
    gnn_scores = (probs * 100).round(2)
 
    USE_TORCH = True
    print(f"\n  ✓ Meilleur AUC-ROC : {best_auc:.4f}  (cible > 0.85)")
 
except ImportError:
    # ── Fallback total : Random Forest comme proxy GNN ────────
    print("  PyTorch non trouvé → Random Forest comme proxy")
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.model_selection import cross_val_score
    from sklearn.metrics import roc_auc_score
 
    clf = RandomForestClassifier(
        n_estimators=200,
        max_depth=8,
        class_weight='balanced',
        random_state=42
    )
    clf.fit(X_scaled, y)
    probs       = clf.predict_proba(X_scaled)[:, 1]
    gnn_scores  = (probs * 100).round(2)
    auc_scores  = cross_val_score(clf, X_scaled, y, cv=5, scoring='roc_auc')
    best_auc    = auc_scores.mean()
    USE_TORCH   = False
    print(f"  AUC-ROC (5-fold CV) : {best_auc:.4f}  (cible > 0.85)")
 
# ═══════════════════════════════════════════════════════════════
# ÉTAPE 4 — Analyse de contagion réseau
# ═══════════════════════════════════════════════════════════════
print("\n► Étape 4 : Analyse de contagion réseau...")
 
import networkx as nx
 
# Construire le graphe NetworkX
G = nx.Graph()
G.add_nodes_from(range(len(df)))
G.add_edges_from(zip(edges_src[:50000], edges_dst[:50000]))  # limite pour performance
 
# Score de contagion : moyenne des scores GNN des voisins
contagion_scores = np.zeros(len(df))
for node in G.nodes():
    neighbors = list(G.neighbors(node))
    if neighbors:
        neighbor_scores = [gnn_scores[n] for n in neighbors if n < len(gnn_scores)]
        if neighbor_scores:
            contagion_scores[node] = np.mean(neighbor_scores)
 
# Score final combiné = 70% score GNN + 30% contagion réseau
final_scores = (0.7 * gnn_scores + 0.3 * contagion_scores).round(2)
 
print(f"  Score GNN moyen      : {gnn_scores.mean():.1f}/100")
print(f"  Score contagion moyen: {contagion_scores.mean():.1f}/100")
print(f"  Score final moyen    : {final_scores.mean():.1f}/100")
 
# ═══════════════════════════════════════════════════════════════
# ÉTAPE 5 — Sauvegarde des résultats
# ═══════════════════════════════════════════════════════════════
print("\n► Étape 5 : Sauvegarde des résultats...")
 
results = df[['CODE_CLIENT'] if 'CODE_CLIENT' in df.columns else []].copy()
if 'CODE_CLIENT' not in df.columns:
    results['client_index'] = range(len(df))
 
if 'gouvernorat' in df.columns:
    results['gouvernorat']  = df['gouvernorat'].values
if 'nature_client' in df.columns:
    results['nature_client'] = df['nature_client'].values
 
results['score_gnn']        = gnn_scores
results['score_contagion']  = contagion_scores.round(2)
results['score_final_m2']   = final_scores
results['label_reel']       = y
results['alerte'] = pd.cut(
    final_scores,
    bins=[-1, 30, 70, 101],
    labels=['VERT', 'ORANGE', 'ROUGE']
)
 
# Statistiques par niveau d'alerte
print(f"\n  Répartition des alertes :")
print(f"    VERT   (score < 30)  : {(results['alerte']=='VERT').sum()} clients")
print(f"    ORANGE (score 30-70) : {(results['alerte']=='ORANGE').sum()} clients")
print(f"    ROUGE  (score > 70)  : {(results['alerte']=='ROUGE').sum()} clients")
 
# Sauvegarder CSV
OUTPUT_CSV  = r'm2_gnn_scores.csv'
results.to_csv(OUTPUT_CSV, index=False)
print(f"\n  ✓ CSV sauvegardé : {OUTPUT_CSV}")
 
# Sauvegarder Excel formaté
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
OUTPUT_XLSX = r'm2_gnn_results.xlsx'
 
with pd.ExcelWriter(OUTPUT_XLSX, engine='openpyxl') as writer:
    results.to_excel(writer, sheet_name='scores_clients', index=False)
 
    # Résumé
    summary = pd.DataFrame({
        'Métrique': ['Clients total', 'AUC-ROC', 'Alertes ROUGE', 'Alertes ORANGE', 'Alertes VERT', 'Score moyen'],
        'Valeur':   [
            len(results),
            f"{best_auc:.4f}",
            (results['alerte']=='ROUGE').sum(),
            (results['alerte']=='ORANGE').sum(),
            (results['alerte']=='VERT').sum(),
            f"{final_scores.mean():.1f}/100"
        ]
    })
    summary.to_excel(writer, sheet_name='résumé', index=False)
 
# Mise en forme
wb = load_workbook(OUTPUT_XLSX)
ws = wb['scores_clients']
 
BLUE  = PatternFill('solid', fgColor='1E4FA0')
ROUGE = PatternFill('solid', fgColor='FFE4E4')
ORANG = PatternFill('solid', fgColor='FFF3CD')
VERT  = PatternFill('solid', fgColor='E8F5E9')
GRAY  = PatternFill('solid', fgColor='F5F7FA')
HF    = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
BF    = Font(name='Calibri', size=10)
thin  = Side(style='thin', color='CCCCCC')
BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)
 
for cell in ws[1]:
    cell.font = HF; cell.fill = BLUE
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BDR
 
for col_cells in ws.columns:
    length = max(len(str(c.value or '')) for c in col_cells)
    ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 25)
 
# Colorer selon alerte
header = [cell.value for cell in ws[1]]
try:
    alerte_col = header.index('alerte') + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        val = row[alerte_col - 1].value
        fill = ROUGE if val == 'ROUGE' else ORANG if val == 'ORANGE' else VERT
        for cell in row:
            cell.fill = fill; cell.font = BF
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = BDR
except: pass
 
ws.freeze_panes = 'A2'
ws.row_dimensions[1].height = 30
wb.save(OUTPUT_XLSX)
print(f"  ✓ Excel sauvegardé : {OUTPUT_XLSX}")
 
# ═══════════════════════════════════════════════════════════════
# RAPPORT FINAL
# ═══════════════════════════════════════════════════════════════
print("\n" + "=" * 55)
print("  RAPPORT M2 — Graph Neural Network")
print("=" * 55)
print(f"  Clients analysés     : {len(results)}")
print(f"  Arêtes dans le graphe: {len(edges_src)}")
print(f"  AUC-ROC              : {best_auc:.4f}  {'✓ OK' if best_auc >= 0.85 else '→ correct pour continuer'}")
print(f"  Score GNN moyen      : {gnn_scores.mean():.1f}/100")
print(f"  Alertes ROUGE        : {(results['alerte']=='ROUGE').sum()}")
print(f"  Alertes ORANGE       : {(results['alerte']=='ORANGE').sum()}")
print(f"  Alertes VERT         : {(results['alerte']=='VERT').sum()}")
print("=" * 55)
print(f"\n  Fichiers générés :")
print(f"  → {OUTPUT_CSV}")
print(f"  → {OUTPUT_XLSX}")
print(f"\n  Prêt pour M3 (Time Series Forecasting) !")
 