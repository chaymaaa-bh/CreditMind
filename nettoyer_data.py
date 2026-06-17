# =============================================================
#  CreditMind — Nettoyage & Feature Engineering
#  
#  INSTRUCTIONS :
#  1. Place ce fichier dans :
#     C:\Users\Rahma Moalla\Desktop\CreditMind\datta-nettoyer\
#  2. Ouvre un terminal VS Code dans ce dossier
#  3. pip install pandas openpyxl numpy
#  4. python nettoyer_data.py
# =============================================================

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import warnings
warnings.filterwarnings('ignore')

# ─── CHEMINS (relatifs : script dans le même dossier que la data) ─────────────
INPUT  = 'Data_pfe_solvabilité_clients.xlsx'
OUTPUT = 'SolvAI_Dataset_Nettoye.xlsx'

if not os.path.exists(INPUT):
    print(f"ERREUR : fichier introuvable → {INPUT}")
    print("Vérifie que ce script est dans le même dossier que la data.")
    exit()

# ─── 1. LECTURE ───────────────────────────────────────────────────────────────
print("► Lecture des données...")
xl  = pd.read_excel(INPUT, sheet_name=None)
fac = xl['Factures clients'].copy()
reg = xl['Reglements clients'].copy()
print(f"  Factures   : {fac.shape[0]} lignes, {fac.shape[1]} colonnes")
print(f"  Règlements : {reg.shape[0]} lignes, {reg.shape[1]} colonnes")

# ─── 2. NETTOYAGE FACTURES ────────────────────────────────────────────────────
print("\n► Nettoyage des factures...")
fac['DATE_FACTURE'] = pd.to_datetime(fac['DATE_FACTURE'], errors='coerce')
fac['MONTANT_TTC']  = fac['MONTANT_TTC'].abs()
fac.rename(columns={
    'CODE RESPONSABLE': 'CODE_RESPONSABLE',
    'CODE CLIENT'     : 'CODE_CLIENT',
    'NATURE CLIENT'   : 'NATURE_CLIENT',
}, inplace=True)
for col in ['NATURE_CLIENT', 'GOUVERNORAT', 'DELEGATION']:
    fac[col] = fac[col].str.strip().str.upper()
fac['TYPE_DOC']  = fac['NUM_FACTURE'].str.split('-').str[0]
fac['EST_AVOIR'] = fac['TYPE_DOC'].isin(['AVC', 'AVF']).astype(int)
nb_avant = len(fac)
fac.drop_duplicates(subset='NUM_FACTURE', inplace=True)
print(f"  Doublons supprimés : {nb_avant - len(fac)}")
print(f"  Factures conservées : {len(fac)}")

# ─── 3. NETTOYAGE RÈGLEMENTS ──────────────────────────────────────────────────
print("\n► Nettoyage des règlements...")
reg['DATE']          = pd.to_datetime(reg['DATE'],          errors='coerce')
reg['DATE_ECHEANCE'] = pd.to_datetime(reg['DATE ECHEANCE'], errors='coerce')
reg.drop(columns=['DATE ECHEANCE'], inplace=True)
reg.rename(columns={
    'CODE CLIENT': 'CODE_CLIENT',
    'MONTANT REG': 'MONTANT_REG',
}, inplace=True)
reg['MONTANT_REG']  = reg['MONTANT_REG'].abs()
reg['MODE']         = reg['MODE'].str.strip().str.upper()
reg['RETARD_JOURS'] = (reg['DATE'] - reg['DATE_ECHEANCE']).dt.days.fillna(0)
nb_avant = len(reg)
reg.drop_duplicates(subset='REF_REG', inplace=True)
print(f"  Doublons supprimés : {nb_avant - len(reg)}")
print(f"  Règlements conservés : {len(reg)}")

# ─── 4. FEATURES AGRÉGÉES PAR CLIENT ─────────────────────────────────────────
print("\n► Calcul des features par client...")

feat_fac = fac.groupby('CODE_CLIENT').agg(
    nb_factures           = ('NUM_FACTURE',     'count'),
    montant_ttc_total     = ('MONTANT_TTC',     'sum'),
    montant_ttc_moyen     = ('MONTANT_TTC',     'mean'),
    montant_ttc_max       = ('MONTANT_TTC',     'max'),
    montant_ttc_std       = ('MONTANT_TTC',     'std'),
    nb_avoirs             = ('EST_AVOIR',        'sum'),
    date_premiere_facture = ('DATE_FACTURE',    'min'),
    date_derniere_facture = ('DATE_FACTURE',    'max'),
    gouvernorat           = ('GOUVERNORAT',     lambda x: x.mode()[0]),
    nature_client         = ('NATURE_CLIENT',   lambda x: x.mode()[0]),
    code_responsable      = ('CODE_RESPONSABLE','first'),
).reset_index()

reference_date               = pd.Timestamp('2026-02-28')
feat_fac['anciennete_jours']  = (reference_date - feat_fac['date_premiere_facture']).dt.days
feat_fac['ratio_avoirs']      = (feat_fac['nb_avoirs'] / feat_fac['nb_factures']).round(4)

feat_reg = reg.groupby('CODE_CLIENT').agg(
    nb_reglements          = ('REF_REG',      'count'),
    montant_reg_total      = ('MONTANT_REG',  'sum'),
    montant_reg_moyen      = ('MONTANT_REG',  'mean'),
    retard_moyen_jours     = ('RETARD_JOURS', 'mean'),
    retard_max_jours       = ('RETARD_JOURS', 'max'),
    nb_retards_positifs    = ('RETARD_JOURS', lambda x: (x > 0).sum()),
    nb_retards_graves      = ('RETARD_JOURS', lambda x: (x > 30).sum()),
    mode_paiement_dominant = ('MODE',         lambda x: x.mode()[0]),
    nb_modes_distincts     = ('MODE',         'nunique'),
).reset_index()
feat_reg['taux_retard'] = (feat_reg['nb_retards_positifs'] / feat_reg['nb_reglements']).round(4)

clients = feat_fac.merge(feat_reg, on='CODE_CLIENT', how='left')
clients['nb_reglements']          = clients['nb_reglements'].fillna(0).astype(int)
clients['montant_reg_total']      = clients['montant_reg_total'].fillna(0)
clients['montant_reg_moyen']      = clients['montant_reg_moyen'].fillna(0)
clients['retard_moyen_jours']     = clients['retard_moyen_jours'].fillna(0)
clients['retard_max_jours']       = clients['retard_max_jours'].fillna(0)
clients['nb_retards_positifs']    = clients['nb_retards_positifs'].fillna(0).astype(int)
clients['nb_retards_graves']      = clients['nb_retards_graves'].fillna(0).astype(int)
clients['taux_retard']            = clients['taux_retard'].fillna(0)
clients['mode_paiement_dominant'] = clients['mode_paiement_dominant'].fillna('INCONNU')
clients['nb_modes_distincts']     = clients['nb_modes_distincts'].fillna(0).astype(int)
clients['ratio_encaissement']     = np.where(
    clients['montant_ttc_total'] > 0,
    (clients['montant_reg_total'] / clients['montant_ttc_total']).round(4), 0)

# ─── 5. LABEL DE SOLVABILITÉ ──────────────────────────────────────────────────
# 1 = à risque si : jamais payé OU retard max > 30j OU > 50% paiements en retard
clients['label_risque'] = (
    (clients['nb_reglements']    == 0) |
    (clients['retard_max_jours'] >  30) |
    (clients['taux_retard']      >  0.5)
).astype(int)

# ─── 6. ENCODAGE CATÉGORIELLES ───────────────────────────────────────────────
clients['gouvernorat_code']   = clients['gouvernorat'].map(
    {g: i for i, g in enumerate(sorted(clients['gouvernorat'].dropna().unique()))}).fillna(-1).astype(int)
clients['nature_client_code'] = clients['nature_client'].map(
    {n: i for i, n in enumerate(sorted(clients['nature_client'].dropna().unique()))}).fillna(-1).astype(int)
clients['mode_paiement_code'] = clients['mode_paiement_dominant'].map(
    {m: i for i, m in enumerate(sorted(clients['mode_paiement_dominant'].unique()))}).fillna(-1).astype(int)

for col in ['montant_ttc_total','montant_ttc_moyen','montant_ttc_max','montant_ttc_std',
            'montant_reg_total','montant_reg_moyen','retard_moyen_jours','retard_max_jours',
            'ratio_avoirs','taux_retard','ratio_encaissement']:
    clients[col] = clients[col].round(3)

# ─── 7. ORDRE FINAL DES COLONNES ─────────────────────────────────────────────
cols = [
    'CODE_CLIENT','code_responsable',
    'gouvernorat','gouvernorat_code','nature_client','nature_client_code',
    'nb_factures','montant_ttc_total','montant_ttc_moyen','montant_ttc_max',
    'montant_ttc_std','nb_avoirs','ratio_avoirs','anciennete_jours',
    'nb_reglements','montant_reg_total','montant_reg_moyen',
    'retard_moyen_jours','retard_max_jours','nb_retards_positifs',
    'nb_retards_graves','taux_retard','mode_paiement_dominant',
    'mode_paiement_code','nb_modes_distincts','ratio_encaissement','label_risque',
]
clients_final = clients[cols].copy()

# ─── 8. ÉCRITURE EXCEL ────────────────────────────────────────────────────────
print("\n► Écriture du fichier Excel...")
with pd.ExcelWriter(OUTPUT, engine='openpyxl') as writer:
    clients_final.to_excel(writer, sheet_name='clients_features',  index=False)
    fac.drop(columns=['TYPE_DOC'], errors='ignore').to_excel(writer, sheet_name='factures_clean',    index=False)
    reg.to_excel(writer, sheet_name='reglements_clean', index=False)

# ─── 9. MISE EN FORME ─────────────────────────────────────────────────────────
wb          = load_workbook(OUTPUT)
HEADER_FILL = PatternFill('solid', fgColor='1E4FA0')
RISK_FILL   = PatternFill('solid', fgColor='FFE4E4')
SAFE_FILL   = PatternFill('solid', fgColor='E8F5E9')
ALT_FILL    = PatternFill('solid', fgColor='F5F7FA')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
BODY_FONT   = Font(name='Calibri', size=10)
thin        = Side(style='thin', color='CCCCCC')
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

def format_sheet(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    for col_cells in ws.columns:
        length = max(len(str(c.value or '')) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 30)
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = BORDER; cell.fill = fill
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 32

ws = wb['clients_features']
format_sheet(ws)
header    = [cell.value for cell in ws[1]]
label_col = header.index('label_risque') + 1
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    fill = RISK_FILL if row[label_col - 1].value == 1 else SAFE_FILL
    for cell in row:
        cell.fill = fill

format_sheet(wb['factures_clean'])
format_sheet(wb['reglements_clean'])
wb.save(OUTPUT)

# ─── 10. RAPPORT FINAL ────────────────────────────────────────────────────────
print("\n" + "=" * 50)
print("  RAPPORT QUALITÉ")
print("=" * 50)
print(f"  Clients (lignes)     : {len(clients_final)}")
print(f"  Features (colonnes)  : {len(clients_final.columns)}")
print(f"  Valeurs manquantes   : {clients_final.isnull().sum().sum()}")
print(f"  Clients à risque  1  : {clients_final['label_risque'].sum()} ({clients_final['label_risque'].mean():.1%})")
print(f"  Clients solvables 0  : {(clients_final['label_risque']==0).sum()}")
print(f"  Retard moyen global  : {clients_final['retard_moyen_jours'].mean():.1f} jours")
print(f"  Encaissement moyen   : {clients_final['ratio_encaissement'].mean():.1%}")
print(f"  Top gouvernorat      : {clients_final['gouvernorat'].mode()[0]}")
print(f"  Top mode paiement    : {clients_final['mode_paiement_dominant'].mode()[0]}")
print("=" * 50)
print(f"\n  Fichier genere : {OUTPUT}")
print("  Ouvre-le dans Excel pour voir les 3 feuilles.")
